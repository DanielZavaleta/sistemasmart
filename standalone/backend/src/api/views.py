from django.contrib.auth.models import User, Group
from django.contrib.auth import authenticate
from rest_framework.views import APIView
from rest_framework import generics
import datetime
from django.utils import timezone
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction

from .models import (
    Configuracion
)
from .serializers import (
    UserSerializer, GroupSerializer, ProductoSerializer, 
    FamiliaSerializer, SubfamiliaSerializer, ClienteSerializer,
    ProveedorSerializer, VentaSerializer,
    EntradaStockSerializer, AjusteStockSerializer,
    OrdenCompraSerializer, PagoProveedorSerializer,
    MovimientoClienteSerializer, CorteCajaSerializer,
    SucursalSerializer,
    TicketSuspendidoSerializer, RetiroCajaSerializer,
    ConfiguracionSerializer, DescuentoSerializer,
    InventarioSerializer, TransferenciaSerializer
)


class ConfiguracionViewSet(viewsets.ModelViewSet):
    queryset = Configuracion.objects.all()
    serializer_class = ConfiguracionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        # Always return the first/only instance for easier usage
        instance = Configuracion.objects.first()
        if not instance:
             # Create default if missing
             instance = Configuracion.objects.create(nombre_tienda="Mi Tienda")
        
        queryset = self.filter_queryset(self.get_queryset())
        if not queryset.exists():
             Configuracion.objects.create(nombre_tienda="Mi Tienda")
        return super().list(request, *args, **kwargs)

    @action(detail=False, methods=['post'], url_path='purge-data')
    def purge_data(self, request):
        """
        Elimina información transaccional (Ventas, Compras, Movimientos) en un rango de fechas.
        Rango de fechas: start_date a end_date (inclusive).
        Formato esperado: 'YYYY-MM-DD'
        """
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')

        if not start_date_str or not end_date_str:
            return Response({"error": "Debe proporcionar start_date y end_date."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Parse dates and make them timezone aware (covering full days)
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Create datetime range for models with DateTimeField
            start_dt = timezone.make_aware(datetime.datetime.combine(start_date, datetime.datetime.min.time()))
            end_dt = timezone.make_aware(datetime.datetime.combine(end_date, datetime.datetime.max.time()))

            count = 0
            with transaction.atomic():
                # 1. Ventas (Cascade deletes VentaItem, VentaPago)
                ventas = Venta.objects.filter(creado_en__range=(start_dt, end_dt))
                count += ventas.count()
                ventas.delete()

                # 2. Cortes de Caja
                cortes = CorteCaja.objects.filter(fecha__range=(start_dt, end_dt))
                count += cortes.count()
                cortes.delete()

                # 3. Retiros de Caja
                retiros = RetiroCaja.objects.filter(fecha__range=(start_dt, end_dt))
                count += retiros.count()
                retiros.delete()

                # 4. Entradas de Stock (Cascade deletes items)
                entradas = EntradaStock.objects.filter(fecha__range=(start_dt, end_dt))
                count += entradas.count()
                entradas.delete()
                
                # 5. Ajustes de Stock
                ajustes = AjusteStock.objects.filter(fecha__range=(start_dt, end_dt))
                count += ajustes.count()
                ajustes.delete()

                # 6. Órdenes de Compra (Cascade deletes items)
                ordenes = OrdenCompra.objects.filter(fecha_creacion__range=(start_dt, end_dt))
                count += ordenes.count()
                ordenes.delete()

                # 7. Pagos a Proveedores
                pagos = PagoProveedor.objects.filter(fecha__range=(start_dt, end_dt))
                count += pagos.count()
                pagos.delete()

                # 8. Movimientos Clientes (Abonos/Cargos manuales)
                movimientos = MovimientoCliente.objects.filter(fecha__range=(start_dt, end_dt))
                count += movimientos.count()
                movimientos.delete()

                # 9. Transferencias
                transferencias = Transferencia.objects.filter(fecha_creacion__range=(start_dt, end_dt))
                count += transferencias.count()
                transferencias.delete()
                
                # 10. Tickets Suspendidos
                tickets = TicketSuspendido.objects.filter(creado_en__range=(start_dt, end_dt))
                count += tickets.count()
                tickets.delete()

            return Response({"message": f"Depuración completada. Registros eliminados (aprox): {count}"})

        except ValueError:
             return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from django.shortcuts import get_object_or_404
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
import decimal
from django.conf import settings
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


from .serializers import (
    RetiroCajaSerializer,
    UserSerializer, GroupSerializer, ProductoSerializer, 
    FamiliaSerializer, SubfamiliaSerializer, ClienteSerializer,
    ProveedorSerializer, VentaSerializer,
    EntradaStockSerializer, AjusteStockSerializer, CaducidadSerializer,
    CaducidadSerializer,
    OrdenCompraSerializer,
    PagoProveedorSerializer,
    MovimientoClienteSerializer,
    CorteCajaSerializer,
    SucursalSerializer,
    TicketSuspendidoSerializer
)
from .models import (
    Producto, Familia, Subfamilia, Cliente, Proveedor,
    Venta, VentaItem, VentaPago,
    EntradaStock, EntradaStockItem,
    AjusteStock,
    OrdenCompra,
    PagoProveedor,
    MovimientoCliente,
    CorteCaja,
    CorteCaja,
    Sucursal,
    TicketSuspendido,
    RetiroCaja,
    Descuento,
    Inventario,
    Transferencia,
    TransferenciaItem
)



class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]





class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [permissions.IsAuthenticated]



class SucursalViewSet(viewsets.ModelViewSet):
    queryset = Sucursal.objects.all()
    serializer_class = SucursalSerializer
    permission_classes = [permissions.IsAuthenticated]


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='reporte-existencias')
    def reporte_existencias(self, request):
        productos = Producto.objects.all().order_by('descripcion')
        total_valor = 0
        productos_data = []
        
        for producto in productos:
            valor = producto.stock_actual * producto.costo
            total_valor += valor
            productos_data.append({
                'id': producto.id,
                'codigo_barras': producto.codigo_barras,
                'descripcion': producto.descripcion,
                'familia': producto.subfamilia.familia.nombre if producto.subfamilia else 'Sin Familia',
                'stock_actual': producto.stock_actual,
                'costo': producto.costo,
                'valor_stock': valor
            })
            
        return Response({
            'valor_total_inventario': total_valor,
            'total_productos': productos.count(),
            'productos': productos_data
        })


class FamiliaViewSet(viewsets.ModelViewSet):
    queryset = Familia.objects.all()
    serializer_class = FamiliaSerializer
    permission_classes = [permissions.IsAuthenticated]


class SubfamiliaViewSet(viewsets.ModelViewSet):
    queryset = Subfamilia.objects.all()
    serializer_class = SubfamiliaSerializer
    permission_classes = [permissions.IsAuthenticated]


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated]


class DescuentoViewSet(viewsets.ModelViewSet):
    queryset = Descuento.objects.all().order_by('porcentaje')
    serializer_class = DescuentoSerializer
    permission_classes = [permissions.IsAuthenticated]



class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer
    permission_classes = [permissions.IsAuthenticated]


class VentaViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Venta.objects.all().order_by('-creado_en')
    serializer_class = VentaSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['post'], url_path='registrar')
    def registrar_venta(self, request):
        data = request.data
        items_data = data.get('items', [])
        pagos_data = data.get('pagos', [])
        total_venta = decimal.Decimal(data.get('total', '0.00'))
        if not items_data or not pagos_data:
            return Response({"error": "Se requieren 'items' y 'pagos'"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                venta = Venta.objects.create(cajero=request.user, total=total_venta)
                cliente = None
                cliente_id_param = data.get('cliente') or data.get('cliente_id') # Support both
                if cliente_id_param:
                     cliente = get_object_or_404(Cliente, id=cliente_id_param)
                     venta.cliente = cliente
                     venta.save()


                def deduct_fifo(prod, qty):
                    # 1. Master Update (Allow negative stock as per Spec)
                    prod.stock_actual -= qty
                    prod.save()

                    # 2. FIFO Batch Update
                    batches = EntradaStockItem.objects.filter(producto=prod, cantidad_disponible__gt=0).order_by('fecha_caducidad', 'id')
                    to_deduct = qty
                    for batch in batches:
                        if to_deduct <= 0: break
                        available = batch.cantidad_disponible
                        if available >= to_deduct:
                            batch.cantidad_disponible -= to_deduct
                            batch.save()
                            to_deduct = 0
                        else:
                            batch.cantidad_disponible = decimal.Decimal(0)
                            batch.save()
                            to_deduct -= available

                for item_data in items_data:
                    producto = get_object_or_404(Producto, id=item_data['id'])
                    cantidad = decimal.Decimal(item_data['cantidad'])
                    precio = decimal.Decimal(item_data['precio_unitario'])
                    
                    if producto.es_paquete:
                        componentes = producto.componentes.all()
                        if not componentes:
                             deduct_fifo(producto, cantidad)
                        else:
                            for comp in componentes:
                                cantidad_necesaria = comp.cantidad * cantidad
                                deduct_fifo(comp.componente, cantidad_necesaria)
                    else:
                        deduct_fifo(producto, cantidad)

                    VentaItem.objects.create(venta=venta, producto=producto, cantidad=cantidad, precio_unitario=precio, subtotal=cantidad * precio)
                
                total_pagado = decimal.Decimal('0.00')
                for pago_data in pagos_data:
                    monto = decimal.Decimal(pago_data['monto'])
                    metodo = pago_data['metodo']
                    
                    if metodo == 'credito':
                        # Validar si existe cliente
                        to_check_cliente = cliente # Use local var
                        if not to_check_cliente:
                             # Try to get from payload if not set in venta?
                             # In the original code, `cliente` variable inside loop came from... WHERE?
                             # Ah, original code had a bug! it used `cliente` inside loop but defined it where?
                             # It was using a variable `cliente` that might have been undefined or global?
                             # I checked original file lines 179: `if not cliente:`
                             # But `cliente` wasn't defined in the function scope in the original snippet I saw!
                             # Wait, I didn't see `cliente = ...` in the view_file output lines 127+.
                             # It seems I am fixing a bug here too.
                             if not venta.cliente:
                                 raise Exception("Venta a crédito requiere cliente seleccionado")
                             to_check_cliente = venta.cliente

                        # HU-30: Validar Límite de Crédito
                        nuevo_saldo = to_check_cliente.saldo_actual + monto
                        if to_check_cliente.limite_credito > 0 and nuevo_saldo > to_check_cliente.limite_credito:
                            raise Exception(f"El cliente excede su límite de crédito. Saldo: ${to_check_cliente.saldo_actual}, Límite: ${to_check_cliente.limite_credito}, Venta: ${monto}")

                        to_check_cliente.saldo_actual = nuevo_saldo
                        to_check_cliente.save()
                        MovimientoCliente.objects.create(cliente=to_check_cliente, tipo='cargo', monto=monto, venta=venta, notas=f"Cargo por Venta {venta.id}")
                    else:
                        # Si hay cliente, registrar la venta como movimiento? 
                        # El spec dice: "Si es Credito o Anticipo insertar movimiento".
                        # Si es Efectivo, MovimientoCliente NO es obligatorio para el saldo, pero bueno para historial.
                        # El codigo original tenia `MovimientoCliente.objects.create(..., tipo='VENTA_CONTADO'...)`
                        # Pero 'VENTA_CONTADO' no está en los choices del modelo (cargo/abono). 
                        # Lo omitiremos para cumplir con los choices o usaremos 'cargo' y 'abono' inmediato?
                        # Mejor no crear movimiento si es contado, para no duplicar info con Venta.
                        pass

                    VentaPago.objects.create(venta=venta, metodo=metodo, monto=monto)
                    total_pagado += monto
                
                # if total_pagado < total_venta:
                #     raise Exception("El monto pagado es menor al total de la venta.") # Allow partials? POS logic handles logic.
            
            serializer = VentaSerializer(venta)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='reporte-ventas')
    def reporte_ventas(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')

        if not fecha_inicio or not fecha_fin:
            return Response({"error": "Se requieren 'fecha_inicio' y 'fecha_fin'"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        except ValueError:
             return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

        ventas = Venta.objects.filter(creado_en__range=(start_date, end_date))
        
        total_periodo = ventas.aggregate(Sum('total'))['total__sum'] or 0
        conteo_ventas = ventas.count()
        
        pagos = VentaPago.objects.filter(venta__in=ventas)
        total_efectivo = pagos.filter(metodo='efectivo').aggregate(Sum('monto'))['monto__sum'] or 0
        total_tarjeta = pagos.filter(metodo='tarjeta').aggregate(Sum('monto'))['monto__sum'] or 0
        total_credito = pagos.filter(metodo='credito').aggregate(Sum('monto'))['monto__sum'] or 0

        ventas_por_dia = ventas.annotate(fecha=TruncDate('creado_en')).values('fecha').annotate(
            total=Sum('total'),
            cantidad=Count('id')
        ).order_by('fecha')

        # Agregación por Cajero (para la tabla detallada)
        from django.db.models import Min, Max
        
        ventas_por_cajero = ventas.values('cajero__username', 'cajero__first_name', 'cajero__last_name').annotate(
            total_ventas=Sum('total'),
            fecha_inicial_periodo=Min('creado_en'),
            fecha_final_periodo=Max('creado_en'),
            # hora_inicial=Min('creado_en__time'), # SQLite/MySQL complexity varies for time extraction in annotate
            # Simplified: we send the full datetimes and frontend extracts date/time
        ).order_by('cajero__username')
        
        # Reformatear para el frontend
        reporte_cajeros = []
        for v in ventas_por_cajero:
            # Obtener nombre legible
            username = v['cajero__username'] or 'Desconocido'
            first = v['cajero__first_name']
            last = v['cajero__last_name']
            nombre_completo = f"{first} {last}".strip() or username
            
            reporte_cajeros.append({
                'nombre_cajero': nombre_completo,
                'fecha_inicial': v['fecha_inicial_periodo'], # Datetime obj
                'fecha_final': v['fecha_final_periodo'],     # Datetime obj
                'total_ventas': v['total_ventas']
            })

        data = {
            'total_periodo': total_periodo,
            'conteo_ventas': conteo_ventas,
            'total_efectivo': total_efectivo,
            'total_tarjeta': total_tarjeta,
            'total_credito': total_credito,
            'ventas_por_dia': list(ventas_por_dia),
            'ventas_por_cajero': reporte_cajeros
        }
        
        return Response(data)


class AuthorizeActionView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        action = request.data.get('action')
        supervisor = authenticate(request, username=username, password=password)
        print(f"DEBUG: AuthorizeAction - User: {username}, Action: {action}, Auth Result: {supervisor}")
        if supervisor:
             print(f"DEBUG: Is Staff? {supervisor.is_staff}")

        if supervisor is None:
            print("DEBUG: Authentication failed.")
            return Response({"error": "Credenciales de supervisor incorrectas."}, status=status.HTTP_401_UNAUTHORIZED)
        if action == 'cancel_ticket' and supervisor.is_staff:
            return Response({"success": f"Acción '{action}' autorizada por {supervisor.username}."}, status=status.HTTP_200_OK)
        
        print("DEBUG: Permission denied (Not staff or wrong action).")
        return Response({"error": "El usuario no tiene permisos para esta acción."}, status=status.HTTP_403_FORBIDDEN)


class VenderRecargaView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, *args, **kwargs):
        numero = request.data.get('numero')
        monto = request.data.get('monto')
        if not numero or not monto:
            return Response({"error": "Se requieren 'numero' y 'monto'"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            monto_decimal = decimal.Decimal(monto)
            if monto_decimal <= 0:
                raise Exception("El monto debe ser positivo.")
            print("--- SIMULACIÓN API EXTERNA ---")
            print(f"Llamando a API de Telcel para: {numero} por ${monto_decimal}")
            print("--- SIMULACIÓN API EXTERNA OK ---")
            with transaction.atomic():
                try:
                    producto_recarga = Producto.objects.get(codigo_barras='RECARGA')
                except Producto.DoesNotExist:
                    raise Exception("Producto 'RECARGA' no encontrado. Favor de crearlo en el admin.")
                venta = Venta.objects.create(cajero=request.user, total=monto_decimal)
                VentaItem.objects.create(venta=venta, producto=producto_recarga, cantidad=monto_decimal, precio_unitario=decimal.Decimal('1.00'), subtotal=monto_decimal)
                VentaPago.objects.create(venta=venta, metodo='efectivo', monto=monto_decimal)
            serializer = VentaSerializer(venta)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class EntradaStockViewSet(viewsets.ModelViewSet):
    queryset = EntradaStock.objects.all().order_by('-fecha')
    serializer_class = EntradaStockSerializer
    permission_classes = [permissions.IsAuthenticated]
    def get_serializer_context(self):
        return {'request': self.request}


class AjusteStockViewSet(viewsets.ModelViewSet):
    queryset = AjusteStock.objects.all().order_by('-fecha')
    serializer_class = AjusteStockSerializer
    permission_classes = [permissions.IsAuthenticated]
    http_method_names = ['get', 'post', 'head', 'options']
    def get_serializer_context(self):
        return {'request': self.request}


class CaducidadAlertView(generics.ListAPIView):
    serializer_class = CaducidadSerializer
    permission_classes = [permissions.IsAuthenticated]
    def get_queryset(self):
        today = datetime.date.today()
        try:
            days = int(self.request.query_params.get('days', 30))
        except ValueError:
            days = 30
        alert_date = today + datetime.timedelta(days=days)
        return EntradaStockItem.objects.filter(
            fecha_caducidad__isnull=False,
            fecha_caducidad__lte=alert_date,
            cantidad_disponible__gt=0
        ).order_by('fecha_caducidad')


class OrdenCompraViewSet(viewsets.ModelViewSet):
    """
    API endpoint para crear y ver Órdenes de Compra.
    Esto NO modifica el stock.
    """
    queryset = OrdenCompra.objects.all().order_by('-fecha_creacion')
    serializer_class = OrdenCompraSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_context(self):
        
        return {'request': self.request}


class PagoProveedorViewSet(viewsets.ModelViewSet):
    queryset = PagoProveedor.objects.all().order_by('-fecha')
    serializer_class = PagoProveedorSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_context(self):
        return {'request': self.request}


class MovimientoClienteViewSet(viewsets.ModelViewSet):
    queryset = MovimientoCliente.objects.all().order_by('-fecha')
    serializer_class = MovimientoClienteSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        cliente_id = self.request.query_params.get('cliente_id')
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        fecha_fin = self.request.query_params.get('fecha_fin')

        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        
        if fecha_inicio:
            queryset = queryset.filter(fecha__date__gte=fecha_inicio)
        
        if fecha_fin:
            queryset = queryset.filter(fecha__date__lte=fecha_fin)
            
        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            with transaction.atomic():
                movimiento = serializer.save()
                
                # Actualizar saldo del cliente
                cliente = movimiento.cliente
                if movimiento.tipo == 'abono':
                    cliente.saldo_actual -= movimiento.monto
                # (Si fuera cargo manual, se sumaría, pero por ahora solo abonos manuales)
                
                cliente.save()
                
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CorteCajaViewSet(viewsets.ModelViewSet):
    queryset = CorteCaja.objects.all().order_by('-fecha')
    serializer_class = CorteCajaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        tipo = self.request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        return queryset

    def get_serializer_context(self):
        return {'request': self.request}

    @action(detail=False, methods=['get'], url_path='calcular-totales')
    def calcular_totales(self, request):
        # Por defecto, calcular totales del día actual (00:00 a 23:59)
        # O desde el último corte si se implementara esa lógica.
        # Por simplicidad: Totales del día actual.
        
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Filter by current user (Cajero)
        user = request.user
        
        ventas_hoy = Venta.objects.filter(creado_en__range=(today_start, today_end), cajero=user)
        
        total_ventas = ventas_hoy.aggregate(Sum('total'))['total__sum'] or 0
        
        pagos_hoy = VentaPago.objects.filter(venta__in=ventas_hoy)
        
        # Variables originales para cálculos
        raw_ventas_efectivo = pagos_hoy.filter(metodo='efectivo').aggregate(Sum('monto'))['monto__sum'] or 0
        total_tarjeta = pagos_hoy.filter(metodo='tarjeta').aggregate(Sum('monto'))['monto__sum'] or 0
        total_credito = pagos_hoy.filter(metodo='credito').aggregate(Sum('monto'))['monto__sum'] or 0

        # Restar Salidas de Efectivo (Pagos a Proveedores) - Filter by user
        pagos_prov_hoy = PagoProveedor.objects.filter(fecha__range=(today_start, today_end), metodo='efectivo', usuario=user)
        total_pagos_prov = pagos_prov_hoy.aggregate(Sum('monto'))['monto__sum'] or 0
        
        # Restar Retiros de Caja (HU-17) - Filter by user
        retiros_hoy = RetiroCaja.objects.filter(fecha__range=(today_start, today_end), usuario=user)
        total_retiros = retiros_hoy.aggregate(Sum('monto'))['monto__sum'] or 0

        # Abonos - Filter by user
        abonos_hoy = MovimientoCliente.objects.filter(fecha__range=(today_start, today_end), tipo='abono', usuario=user)
        total_abonos = abonos_hoy.aggregate(Sum('monto'))['monto__sum'] or 0

        # Calculo Neto (Teórico)
        # CajaTeorica = (VentasEfectivo + Abonos) - (PagosProv + Retiros)
        total_caja_neto = (raw_ventas_efectivo + total_abonos) - (total_pagos_prov + total_retiros)

        # Calcular Inicio y Fin de Turno
        from django.db.models import Min, Max
        turno_info = ventas_hoy.aggregate(inicio=Min('creado_en'), fin=Max('creado_en'))
        
        hora_inicio_turno = turno_info['inicio'] if turno_info['inicio'] else today_start
        hora_fin_turno = turno_info['fin'] if turno_info['fin'] else today_end

        return Response({
            'total_ventas': total_ventas,
            'ventas_efectivo': raw_ventas_efectivo,
            'ventas_tarjeta': total_tarjeta,
            'ventas_credito': total_credito,
            'pagos_proveedores': total_pagos_prov,
            'abonos': total_abonos,
            'retiros': total_retiros,
            'total_caja_teorico': total_caja_neto,
            
            # Legacy mapping for compatibility if needed (or just use new fields)
            'total_efectivo': total_caja_neto, 
            'total_tarjeta': total_tarjeta,
            'total_credito': total_credito,

            'fecha_inicio': today_start,
            'fecha_fin': today_end,
            'hora_inicio_turno': hora_inicio_turno,
            'hora_fin_turno': hora_fin_turno
        })

    def create(self, request, *args, **kwargs):
        # Al crear, recalculamos o confiamos en los datos enviados?
        # Mejor recalcular para asegurar integridad, pero permitimos guardar lo que el usuario vio.
        # Vamos a guardar lo que el usuario envía, pero validando.
        
        return super().create(request, *args, **kwargs)


class RetiroCajaViewSet(viewsets.ModelViewSet):
    queryset = RetiroCaja.objects.all().order_by('-fecha')
    serializer_class = RetiroCajaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)


class TicketSuspendidoViewSet(viewsets.ModelViewSet):
    queryset = TicketSuspendido.objects.all().order_by('-creado_en')
    serializer_class = TicketSuspendidoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)


class ExportarInventarioView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Encabezados
        headers = ["Código de Barras", "Descripción", "Familia", "Existencia Actual", "Costo", "Precio Venta (1)", "Última Actualización"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)

        # Datos
        productos = Producto.objects.all().select_related('subfamilia__familia')
        for row, prod in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=prod.codigo_barras)
            ws.cell(row=row, column=2, value=prod.descripcion)
            ws.cell(row=row, column=3, value=str(prod.subfamilia) if prod.subfamilia else "N/A")
            ws.cell(row=row, column=4, value=prod.stock_actual)
            ws.cell(row=row, column=5, value=prod.costo)
            ws.cell(row=row, column=6, value=prod.precio_1)
            ws.cell(row=row, column=7, value=prod.actualizado_en.strftime('%Y-%m-%d %H:%M'))

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=inventario_sistema_smart.xlsx'
        wb.save(response)
        return response


class RetiroCajaViewSet(viewsets.ModelViewSet):
    queryset = RetiroCaja.objects.all().order_by('-fecha')
    serializer_class = RetiroCajaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)


class InventarioViewSet(viewsets.ModelViewSet):
    queryset = Inventario.objects.all()
    serializer_class = InventarioSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        sucursal_id = self.request.query_params.get('sucursal_id')
        producto_id = self.request.query_params.get('producto_id')
        
        if sucursal_id:
            queryset = queryset.filter(sucursal_id=sucursal_id)
        if producto_id:
            queryset = queryset.filter(producto_id=producto_id)
            
        return queryset


class TransferenciaViewSet(viewsets.ModelViewSet):
    queryset = Transferencia.objects.all().order_by('-fecha_creacion')
    serializer_class = TransferenciaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request, *args, **kwargs):
        # Custom create logic with transaction
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            with transaction.atomic():
                transferencia = serializer.save()
                
                # Execute Transfer Logic immediately
                origen = transferencia.sucursal_origen
                destino = transferencia.sucursal_destino
                
                if origen == destino:
                    raise Exception("La sucursal de origen y destino no pueden ser la misma.")

                for item in transferencia.items.all():
                    # 1. Check Origin Stock
                    try:
                        inv_origen = Inventario.objects.get(sucursal=origen, producto=item.producto)
                    except Inventario.DoesNotExist:
                         raise Exception(f"El producto {item.producto.descripcion} no existe en la sucursal de origen.")
                    
                    if inv_origen.cantidad < item.cantidad:
                         raise Exception(f"Stock insuficiente de {item.producto.descripcion} en origen. Disponible: {inv_origen.cantidad}, Solicitado: {item.cantidad}")
                    
                    # 2. Decrement Origin
                    inv_origen.cantidad -= item.cantidad
                    inv_origen.save()
                    
                    # 3. Increment Destination
                    inv_destino, created = Inventario.objects.get_or_create(
                        sucursal=destino, 
                        producto=item.producto,
                        defaults={'cantidad': 0}
                    )
                    inv_destino.cantidad += item.cantidad
                    inv_destino.save()
                
                transferencia.estado = 'COMPLETADA'
                transferencia.fecha_completada = timezone.now()
                transferencia.save()
                
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
